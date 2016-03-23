"""Excel reader module for a unamed data loader.
Working as a glue between the loader and `xlrd`
and `openpyxl` libraries.

For now, the reader only returns the first sheet,
ignoring the other sheets.

@author: Ares Ou (aresowj@gmail.com)
@version: 20160323
"""

import xlrd
import openpyxl
from xlrd.sheet import XL_CELL_DATE
from openpyxl.utils import get_column_letter


class ExcelReaderBase(object):
    """Base class for excel reader. We store
    book, current sheet and the rows count and
    column count for the sheet in the class
    attributes.
    """
    def __init__(self, file_path):
        self.book = None
        self.sheet = None
        self.rows = 0
        self.columns = 0
        self.index_offset = 0
        self.current_row = 0
        self.file_path = file_path
        self.open_book()

    def __iter__(self):
        return self

    def next(self):
        row = self.current_row
        if row >= self.rows + self.index_offset:
            raise StopIteration
        self.current_row += 1
        return [self.return_cell_value(row, column)
                for column in range(self.index_offset, self.columns + self.index_offset)]

    def open_book(self):
        raise NotImplementedError

    def return_cell_value(self, row, column):
        """Date cells are paid special attention.
        Date parser in data loader is expecting
        a date string from the file, we should
        convert it into string before returning
        the cell value if it is a date cell.
        """
        raise NotImplementedError


class XlsExcelReader(ExcelReaderBase):
    # For extracting year / month / day from the date tuple
    XLS_DATE_RANGE = 3

    def open_book(self):
        self.book = xlrd.open_workbook(self.file_path)
        self.sheet = self.book.sheet_by_index(0)
        self.rows = self.sheet.nrows
        self.columns = self.sheet.ncols

    def return_cell_value(self, row, column):
        cell = self.sheet.cell(row, column)
        if cell.ctype == XL_CELL_DATE:
            return '/'.join(str(d) for d in xlrd.xldate_as_tuple(cell.value, self.book.datemode)[:self.XLS_DATE_RANGE])
        return cell.value


class XlsxExcelReader(ExcelReaderBase):
    def __init__(self, file_path):
        super(XlsxExcelReader, self).__init__(file_path)
        # index of openpyxl starts from 1 but not 0
        self.current_row = 1
        self.index_offset = 1

    def open_book(self):
        self.book = openpyxl.load_workbook(self.file_path)
        self.sheet = self.book.active
        self.rows = self.sheet.max_row
        self.columns = self.sheet.max_column

    def return_cell_value(self, row, column):
        """Note: openpyxl is accessing the cell with
        the excel verbose format `A1`. We use the utility
        function from the library to get column letter
        for every column.
        """
        cell = self.sheet[get_column_letter(column) + str(row)]
        if cell.is_date:
            # when the cell is date format, openpyxl will return a Python datetime object
            return cell.value.strftime('%Y%m%d')
        return cell.value


def get_excel_reader(file_path, extension):
    """Entrance exposed to the data loader"""
    if extension == 'xls':
        return XlsExcelReader(file_path)
    elif extension == 'xlsx':
        return XlsxExcelReader(file_path)
    else:
        return None
